from cmschina_tianyan.session import CmsSession
import pandas as pd

from WindPy import *
#Wind API 使用W账户密码登录
w.start() 
#return True and then can go to next operation.
w.isconnected()

########################################################################################################################
# get ann
dataset = CmsSession.init(appId="cb632213cf764a019b32f8cc9d2040b7", appSecret="b23700b9fbf0090709c7b72ac0fdcb390cd19cc663bcf1c7a07c12b8c32bf377")

stk_code_list = ['300587.SZ',
'603185.SH',
'600674.SH',
'002203.SZ',
'601127.SH',
'002597.SZ',
'603348.SH',
'300487.SZ',
'603876.SH',
'601677.SH',
'300655.SZ',
'002078.SZ',
'002738.SZ', #2022-10-14新增4个对子
'002812.SZ',
'603596.SH',
'300332.SZ',]

# stk_code = '603355.SH'
df_ann_res = []
for stk_code in stk_code_list:
    sql = f''' 
    SELECT * FROM wind_admin.ASHAREANNINF 
    where S_INFO_WINDCODE='{stk_code}' 
    order by ANN_DT desc
    '''

    df_ann = dataset.get_data(sql)
    df_ann_res.append(df_ann)
    # df_ann.to_excel(f'./ann/{stk_code}.xlsx')
df_total_ann = pd.concat(df_ann_res)
df_total_ann['serial'] = [i for i in range(len(df_total_ann))]
df_total_ann['author'] = ['上市公司']*len(df_total_ann)
df_total_ann = df_total_ann.set_index('serial')
df_total_ann = df_total_ann[['S_INFO_WINDCODE','ANN_DT','N_INFO_TITLE','author']]
cols = ['code','date','title','author']
df_total_ann.columns = cols
df_total_ann    
print('*'*50,'公告提取完成','*'*50)
########################################################################################################################
# get news
import pandas as pd
import requests
import random
import time
import os
import openpyxl
from bs4 import BeautifulSoup
def get_url(code,pages):
    '''
    获取东方财富网股吧链接列表
    code是指公司代码
    page是值爬取页数
    '''
    url_list = []
    for page in range(1,pages+1):
        url = f"http://guba.eastmoney.com/list,{code},1,f_{page}.html"
        url_list.append(url)
        
    return url_list

def get_news(url_list,code):
    '''
    获取东方财富网新闻列表至本地xls
    url_list是指链接列表
    '''
    headers = {
        # 'User-Agent': UserAgent(verify_ssl=False).random,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36',
        'cookie': 'qgqp_b_id=b90ae58b50ba4b8868c1988b80823e4c; st_si=62771611099798; st_asi=delete; st_pvi=50748890314615; st_sp=2022-10-24 11:22:01; st_inirUrl=http://guba.eastmoney.com/list,hk01810,1,f_1.html; st_sn=2; st_psi=20221024112234183-117001300541-9748756866'
    }
    
    # 保存爬取内容
    outwb = openpyxl.Workbook() # 打开一个将写的文件
    outws = outwb.create_sheet(index=0) # 在将写的文件创建sheet
    outws.cell(row = 1, column = 1, value = "read")
    outws.cell(row = 1, column = 2, value = "comment")
    outws.cell(row = 1, column = 3, value = "title")
    outws.cell(row = 1, column = 4, value = "author")
    outws.cell(row = 1, column = 5, value = "renew")
    outws.cell(row = 1, column = 6, value = "link")
    index = 2
    
    for i in range(len(url_list)):
        url = url_list[i]
        res = requests.get(url,headers = headers)
        res.encoding = res.apparent_encoding
        html = res.text
        soup = BeautifulSoup(html,"html.parser")
        read_list = soup.select(".l1.a1")[1:]
        comment_list = soup.select(".l2.a2")[1:]
        title_list = soup.select(".l3.a3")[1:]
        author_list = soup.select(".l4.a4")[1:]
        renew_list = soup.select(".l5.a5")[1:]
        for k in range(len(title_list)):
            outws.cell(row = index, column = 1, value = str(read_list[k].text.strip()))
            outws.cell(row = index, column = 2, value = str(comment_list[k].text.strip()))
            outws.cell(row = index, column = 3, value = str(title_list[k].select('a')[0]["title"]))
            outws.cell(row = index, column = 4, value = str(author_list[k].text.strip()))
            outws.cell(row = index, column = 5, value = str(renew_list[k].text.strip()))
            outws.cell(row = index, column = 6, value = str(title_list[k].select('a')[0]["href"]))                                                
            index += 1
            # print(str(title_list[k].select('a')[0]["href"]))
            # print(title_list[k].select('a')[0]["title"],renew_list[k].text.strip())
        time.sleep(random.uniform(3,4))
    save_path = f"./news_ann/news/tmp/{code}.xlsx"            
    outwb.save(save_path)
    df = pd.read_excel(save_path)
    return df 

########################################################################################################################
stk_code_list = ['300587.SZ',
'603185.SH',
'600674.SH',
'002203.SZ',
'601127.SH',
'002597.SZ',
'603348.SH',
'300487.SZ',
'603876.SH',
'601677.SH',
'300655.SZ',
'002078.SZ',
'002738.SZ', #2022-10-14新增4个对子
'002812.SZ',
'603596.SH',
'300332.SZ',]
df_news_res = []
print('*'*50,'开始爬取新闻数据','*'*50)
for stk_code in stk_code_list:
    code = stk_code.split('.')[0]
    pages = 1
    url_list = get_url(code,pages)
    df_tmp = get_news(url_list,code)
    df_tmp['code'] = [stk_code]*len(df_tmp)
    df_news_res.append(df_tmp)
    print(f"{code} 运行完成")    
df_total_news = pd.concat(df_news_res)


df_total_news = pd.concat(df_news_res)
df_total_news['serial'] = [i for i in range(len(df_total_news))]
df_total_news = df_total_news.set_index('serial')

df_total_news['date'] = df_total_news['renew'].apply(lambda x:'2022-' + x.split(' ')[0])
cols = ['code','date','title','author']
df_total_news = df_total_news[cols]
df_total_news
########################################################################################################################
# merge ann and news 
df_total_news_ann = pd.concat([df_total_ann,df_total_news])
df_total_news_ann['serial'] = [i for i in range(len(df_total_news_ann))]
df_total_news_ann = df_total_news_ann.set_index('serial')


df_stk_name = w.wss('300587.SZ,603185.SH,600674.SH,002203.SZ,601127.SH,002597.SZ,603348.SH,300487.SZ,603876.SH,601677.SH,300655.SZ,002078.SZ,002738.SZ, 002812.SZ,603596.SH,300332.SZ', "sec_name", "",usedf=True)[1]
stk_name_dic = df_stk_name.to_dict()['SEC_NAME']
# df_stk_name
df_total_news_ann['name'] =  df_total_news_ann['code'].map(stk_name_dic)
ordered_cols = ['code','name','date','title','author']

df_total_news_ann = df_total_news_ann [ordered_cols]
df_total_news_ann['code_name'] = df_total_news_ann['code'] + df_total_news_ann['name']
df_total_news_ann

df_total_news_ann.to_excel('./news_ann/total_news_ann/news_ann.xlsx')
print('*'*50,'新闻&公告提取完成','*'*50)